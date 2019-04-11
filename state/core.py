import pickle

import openpyxl

from typing import Any, Dict, List, Optional, Tuple
from mypy_extensions import TypedDict

import state.samplesheet as samplesheet
import state.miseq as miseq

from common import label_to_key


_Snapshot = TypedDict('_Snapshot', {
    'samplesheet': samplesheet.SampleSheet,
    'miseq': miseq.MiSeqOutput,
    'ko_mapping': Dict[str, Optional[str]],
    'default_ko_mapping': Dict[str, Optional[str]],
})


_SnapshotList = List[_Snapshot]


KOMapping = Dict[str, Optional[str]]


Clone = TypedDict('Clone', {
    'label': str,
    'comment': Optional[str],
    'knockouts': Dict[str, Optional[miseq.Result]],
})


class State:
    def __init__(self) -> None:
        self.samplesheet = {
            'knockouts': [],
            'headers': [],
            'width': 0,
            'height': 0,
            'group_by': 'id',
        }  # type: samplesheet.SampleSheet

        self.miseq = {}  # type: miseq.MiSeqOutput
        self.ko_mapping = {}  # type: KOMapping
        self.default_ko_mapping = {}  # type: KOMapping

        self._undo_history = []  # type: _SnapshotList
        self._redo_history = []  # type: _SnapshotList
        self._saved_state = None  # type: Optional[_Snapshot]

    def undo(self) -> None:
        if self._undo_history:
            self._redo_history.append(self._store_state())
            self._replace_state(self._undo_history.pop())

    def redo(self) -> None:
        if self._redo_history:
            self._undo_history.append(self._store_state())
            self._replace_state(self._redo_history.pop())

    def samplesheet_load(self, filename: str) -> None:
        state = self._store_state()
        self.samplesheet = samplesheet.load(filename)
        self._init_ko_mapping()
        self._append_undo_history(state)

    def samplesheet_user_split(self, group: str, value: bool) -> None:
        state = self._store_state()
        self.samplesheet = samplesheet.set_user_split(self.samplesheet,
                                                      group=group,
                                                      value=value)
        self._append_undo_history(state)

    def samplesheet_group_by(self, key: str) -> None:
        state = self._store_state()
        self.samplesheet = samplesheet.group_by(self.samplesheet, key)
        self._append_undo_history(state)

    def samplesheet_target_names(self) -> List[str]:
        return sorted({knockout['headers']['knockout']
                       for knockout in self.samplesheet['knockouts']})

    def samplesheet_column(self, knockout: str) -> List[Optional[str]]:
        for key, value in self.ko_mapping.items():
            if value == knockout:
                return samplesheet.get_column_for_ko(self.samplesheet, key)

        return []

    def miseq_load(self, filename: str) -> None:
        state = self._store_state()
        self.miseq = miseq.load(filename)
        self._init_ko_mapping()
        self._append_undo_history(state)

    def miseq_target_names(self) -> List[str]:
        return sorted(self.miseq)

    def miseq_toggle_picked(self, target: str, index: int) -> None:
        state = self._store_state()
        self.miseq = miseq.toggle_picked(self.miseq, target, index)
        self._append_undo_history(state)

    def miseq_set_comment(self, target: str, index: int, comment: str) -> None:
        state = self._store_state()
        self.miseq = miseq.set_comment(self.miseq, target, index, comment)
        self._append_undo_history(state)

    def mapping_set(self, samplesheet: str, miseq: Optional[str]) -> None:
        if self.ko_mapping[samplesheet] != miseq:
            self._append_undo_history(self._store_state())

            new_mapping = {}
            for key, value in self.ko_mapping.items():
                new_mapping[key] = None if value == miseq else value
            new_mapping[samplesheet] = miseq

            self.ko_mapping = new_mapping

    def clones_groups(self) -> List[Tuple[str, bool]]:
        return samplesheet.get_groups(self.samplesheet)

    def clones_get_group(self, group: str) -> List[Clone]:
        clones = {}  # type: Dict[str, Clone]

        for knockout in samplesheet.get_group(self.samplesheet, group):
            ss_target = knockout['headers']['knockout']
            ms_target = self.ko_mapping[ss_target]

            miseq = {} if ms_target is None else self.miseq[ms_target]
            for index, key in enumerate(knockout['column'], start=1):
                if key is not None:
                    try:
                        clone = clones[key]
                    except KeyError:
                        clone = clones[key] = {
                            'label': key,
                            'comment': None,
                            'knockouts': {},
                        }

                    clone['knockouts'][ss_target] = miseq.get(index)

        return sorted(clones.values(),
                      key=lambda clone: label_to_key(clone['label']))

    @property
    def redo_count(self) -> int:
        return len(self._redo_history)

    @property
    def undo_count(self) -> int:
        return len(self._undo_history)

    def export(self, filename: str, everything: bool=False) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.get_active_sheet()

        bd_fat_bottom = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side('medium'))
        bd_thin_bottom = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side('thin'))
        grey_font = openpyxl.styles.Font(color='FFB7B7B7')
        blue_back = openpyxl.styles.PatternFill("solid", fgColor="D3E8EE")

        headers = ('Group', 'Clone', 'KOs', 'Index', 'Reads', 'Indels', '%',
                   '%Total', 'Comment')
        sheet.append(headers)
        for column, _ in enumerate(headers, start=1):
            sheet.cell(row=1, column=column).border = bd_fat_bottom

        last_row = 1
        for group, is_split in self.clones_groups():
            clones = self.clones_get_group(group)
            if not (is_split or everything):
                clones = [clone for clone in clones
                          if any(ko and ko['picked']
                                 for ko in clone['knockouts'].values())]

            if not clones:
                continue

            for column in range(1, len(headers) + 1):
                sheet.cell(row=last_row, column=column).border = bd_fat_bottom

            for not_first_clone, clone in enumerate(clones):
                first_row = True
                if not_first_clone:
                    for column in range(1, len(headers) + 1):
                        sheet.cell(row=last_row, column=column).border = bd_thin_bottom

                for key, knockout in sorted(clone['knockouts'].items()):
                    row = [
                        group if first_row else None,
                        clone['label'] if first_row else None,
                        key,
                    ]  # type: List[Any]

                    if knockout:
                        indels = []
                        indels_pct = []
                        indel_pct_total = 0.0
                        for peak in sorted(knockout['peaks'],
                                           key=lambda peak: peak['pct'],
                                           reverse=True):
                            indels.append('%+i' % (peak['indel'],))
                            indels_pct.append('%i' % (peak['pct'] * 100,))
                            indel_pct_total += peak['pct'] * 100

                        row.extend((
                            int(knockout['index']),
                            knockout['reads'],
                            ' / '.join(indels),
                            ' / '.join(indels_pct),
                            '%i' % (indel_pct_total,),
                            knockout['comment'],
                        ))
                    else:
                        row.extend((
                            '<NA>',
                            '',
                            '',
                            '',
                            '',
                            'Insufficient data',
                        ))

                    sheet.append(row)
                    first_row = False
                    last_row += 1

                    if not (knockout and (is_split or knockout['picked'])):
                        for column in range(3, len(headers) + 1):
                            sheet.cell(row=last_row, column=column).font = grey_font
                    elif knockout['comment'].lower() in ('wt', 'wildtype'):
                        for column in range(1, len(headers) + 1):
                            sheet.cell(row=last_row, column=column).fill = blue_back

            sheet.append([])
            last_row += 1

        workbook.save(filename)

    def load_state(self, filename: str) -> None:
        with open(filename, 'rb') as handle:
            state = self._store_state()
            self._replace_state(pickle.loads(handle.read()))
            self._append_undo_history(state)
        self._saved_state = self._store_state()

    def save_state(self, filename: str) -> None:
        state = self._store_state()
        with open(filename, 'wb') as handle:
            handle.write(pickle.dumps(state))
        self._saved_state = state

    def is_saved(self) -> bool:
        return self._saved_state == self._store_state()

    def _store_state(self) -> _Snapshot:
        return {
            'samplesheet': self.samplesheet,
            'miseq': self.miseq,
            'ko_mapping': self.ko_mapping,
            'default_ko_mapping': self.default_ko_mapping,
        }

    def _append_undo_history(self, state: _Snapshot) -> None:
        self._redo_history.clear()
        self._undo_history.append(state)

    def _replace_state(self, state: _Snapshot) -> None:
        for key, value in state.items():
            setattr(self, key, value)

    def _init_ko_mapping(self) -> None:
        miseq_targets = set(self.miseq_target_names())
        used_miseq_targets = set()
        default_ko_mapping = dict.fromkeys(self.samplesheet_target_names())

        for key in default_ko_mapping:
            if key in miseq_targets:
                default_ko_mapping[key] = key
                used_miseq_targets.add(key)

        for key, value in default_ko_mapping.items():
            if not value:
                candidate = None
                for target in set(miseq_targets) - used_miseq_targets:
                    if target.lower().replace(' ', '_').startswith(key.lower().replace(' ', '_')):
                        if not candidate or len(candidate) > len(target):
                            candidate = target

                if candidate:
                    default_ko_mapping[key] = candidate

        self.default_ko_mapping = default_ko_mapping
        self.ko_mapping = dict(default_ko_mapping)
